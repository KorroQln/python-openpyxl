# Python program to read an excel file
 
# import openpyxl module
import openpyxl
 
# Give the location of the file
path = "C:\\Users\Hazim Aqlan\\Documents\\VS Code\\gitRepo\\python-openpyxl\\Students.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.
 
# Note: The first row or
# column integer is 1, not 0.
 
# Cell object is created by using
# sheet object's cell() method.
cell_obj = sheet_obj.cell(row = 1, column = 1)

# Print value of cell object
# using the value attribute
print(cell_obj.value)

# print the total number of rows/columns
print(sheet_obj.max_row)
print(sheet_obj.max_column)

# Loop will print all columns name
max_col = sheet_obj.max_column
for i in range(1, max_col + 1):
    cell_obj2 = sheet_obj.cell(row = 1, column = i)
    print(cell_obj2.value)
    
# Loop will print all values of first column
m_row = sheet_obj.max_row
for i in range(1, m_row + 1):
    cell_obj3 = sheet_obj.cell(row = i, column = 1)
    print(cell_obj3.value)
    
# Will print a particular row value
max_col2 = sheet_obj.max_column
for i in range(1, max_col2 + 1):
    cell_obj = sheet_obj.cell(row = 2, column = i)
    print(cell_obj.value, end = " ")